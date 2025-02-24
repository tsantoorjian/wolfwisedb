import requests
import pandas as pd
from openpyxl import load_workbook




def fetch_nba_data(url, headers, params):
    response = requests.get(url, headers=headers, params=params)
    data = response.json()
    return data

def process_data(data, min_minutes, min_games):
    headers = data['resultSets'][0]['headers']
    rows = data['resultSets'][0]['rowSet']
    df = pd.DataFrame(rows, columns=headers)
    df_filtered = df[(df['MIN'] >= min_minutes) & (df['GP'] >= min_games)]
    return df_filtered

def rank_players(df, rank_columns):
    for col in rank_columns:
        df[col] = df[col].rank(method='average')
    df.reset_index(drop=True, inplace=True)
    return df


def save_to_excel(df, excel_filename, sheet_name):
    # Load the existing workbook
    workbook = load_workbook(excel_filename)

    # Check if the sheet exists and delete it
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    # Now, use pandas ExcelWriter with the updated workbook
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save the workbook
    workbook.save(excel_filename)


# Assuming other parts of your script have already been defined, such as fetching and processing data

def generate_summary_tab(excel_filename):
    # Load the workbook and check if 'Timberwolves Summary' sheet exists
    wb = load_workbook(excel_filename)
    if 'Timberwolves Summary' in wb.sheetnames:
        del wb['Timberwolves Summary']  # Delete the sheet if it exists
        wb.save(excel_filename)  # Save the workbook after deletion

    sheet_names = wb.sheetnames
    rank_sheets = [name for name in sheet_names if 'Rank' in name and name != 'Timberwolves Summary']

    top_10_ranks_with_timeframe = []

    for sheet in rank_sheets:
        # Specifying header=1 to indicate that the second row contains the column names
        df = pd.read_excel(excel_filename, sheet_name=sheet, header=1)
        # Adjusting based on the actual structure of your sheets
        if 'Player' in df.columns:
            for col in df.columns[2:]:  # Assuming the first two columns are not needed for rank calculation
                df[col] = pd.to_numeric(df[col], errors='coerce')
                top_ranks = df.loc[df[col] <= 10, ['Player', col]]
                for _, row in top_ranks.iterrows():
                    top_10_ranks_with_timeframe.append({
                        "Player": row['Player'],
                        "Stat": col,
                        "Rank": row[col],
                        "Timeframe": sheet
                    })

    # Convert the list to DataFrame, sort, and save
    if top_10_ranks_with_timeframe:
        summary_df = pd.DataFrame(top_10_ranks_with_timeframe).sort_values(by="Player")
        save_to_excel(summary_df, excel_filename, 'Timberwolves Summary')
    else:
        print("No top 10 ranks found.")

# Common headers and URL
url = "https://stats.nba.com/stats/leaguedashplayerstats"
url2 = 'https://stats.nba.com/stats/synergyplaytypes'
headers = {
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9,pt-BR;q=0.8,pt;q=0.7",
    "Connection": "keep-alive",
    "Host": "stats.nba.com",
    "Origin": "https://www.nba.com",
    "Referer": "https://www.nba.com/",
    "Sec-Ch-Ua": "\"Not A Brand\";v=\"99\", \"Google Chrome\";v=\"121\", \"Chromium\";v=\"121\"",
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": "\"macOS\"",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
}
params1 = {
    "College": "",
    "Conference": "",
    "Country": "",
    "DateFrom": "",
    "DateTo": "",
    "Division": "",
    "DraftPick": "",
    "DraftYear": "",
    "GameScope": "",
    "GameSegment": "",
    "Height": "",
    "ISTRound": "",
    "LastNGames": "5",
    "LeagueID": "00",
    "Location": "",
    "MeasureType": "Advanced",
    "Month": "0",
    "OpponentTeamID": "0",
    "Outcome": "",
    "PORound": "0",
    "PaceAdjust": "N",
    "PerMode": "Totals",
    "Period": "0",
    "PlayerExperience": "",
    "PlayerPosition": "",
    "PlusMinus": "N",
    "Rank": "N",
    "Season": "2023-24",
    "SeasonSegment": "",
    "SeasonType": "Regular Season",
    "ShotClockRange": "",
    "StarterBench": "",
    "TeamID": "0",
    "VsConference": "",
    "VsDivision": "",
    "Weight": ""
}
params5 = {
    "LeagueID": "00",
    "PerMode": "Totals",
    "PlayType": "OffScreen",
    "PlayerOrTeam": "P",
    "SeasonType": "Regular Season",
    "SeasonYear": "2023-24",
    "TypeGrouping": "offensive"
}
# Common rank columns
rank_columns = [
    "GP_RANK", "W_RANK", "L_RANK", "W_PCT_RANK", "MIN_RANK",
    "E_OFF_RATING_RANK", "OFF_RATING_RANK", "sp_work_OFF_RATING_RANK",
    "E_DEF_RATING_RANK", "DEF_RATING_RANK", "sp_work_DEF_RATING_RANK",
    "E_NET_RATING_RANK", "NET_RATING_RANK", "sp_work_NET_RATING_RANK",
    "AST_PCT_RANK", "AST_TO_RANK", "AST_RATIO_RANK", "OREB_PCT_RANK",
    "DREB_PCT_RANK", "REB_PCT_RANK", "TM_TOV_PCT_RANK", "E_TOV_PCT_RANK",
    "EFG_PCT_RANK", "TS_PCT_RANK", "USG_PCT_RANK", "E_USG_PCT_RANK",
    "E_PACE_RANK", "PACE_RANK", "sp_work_PACE_RANK", "PIE_RANK",
    "FGM_RANK", "FGA_RANK", "FGM_PG_RANK", "FGA_PG_RANK", "FG_PCT_RANK"
]

rank_columns2 = [
    "FGM_RANK", "FGA_RANK", "FG_PCT_RANK", "FG3M_RANK", "FG3A_RANK",
    "FG3_PCT_RANK", "FTM_RANK", "FTA_RANK", "FT_PCT_RANK", "OREB_RANK",
    "DREB_RANK", "REB_RANK", "AST_RANK", "TOV_RANK", "STL_RANK",
    "BLK_RANK", "BLKA_RANK", "PF_RANK", "PFD_RANK", "PTS_RANK",
    "PLUS_MINUS_RANK", "NBA_FANTASY_PTS_RANK", "DD2_RANK", "TD3_RANK",
    "WNBA_FANTASY_PTS_RANK"
]

rank_columns3 = [
    'SEASON_ID', 'PLAYER_ID', 'PLAYER_NAME', 'TEAM_ID',
    'TEAM_ABBREVIATION', 'TEAM_NAME', 'PLAY_TYPE', 'TYPE_GROUPING',
    'PERCENTILE', 'GP', 'POSS_PCT', 'PPP', 'FG_PCT', 'FT_POSS_PCT',
    'TOV_POSS_PCT', 'SF_POSS_PCT', 'PLUSONE_POSS_PCT', 'SCORE_POSS_PCT',
    'EFG_PCT', 'POSS', 'PTS', 'FGM', 'FGA', 'FGMX'
]

# Parameters for the second request with a small change
params2 = params1.copy()
params2['LastNGames'] = "10"  # Adjusting the last N games parameter

params3 = params1.copy()
params3['MeasureType'] = "Base"  # Adjusting the last N games parameter

params4 = params2.copy()
params4['MeasureType'] = "Base"  # Adjusting the last N games parameter

# Fetch and process the first set of data
data1 = fetch_nba_data(url, headers, params1)
df_filtered1 = process_data(data1, min_minutes=20, min_games=4)
df_ranked1 = rank_players(df_filtered1, rank_columns)
save_to_excel(df_ranked1, '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx', 'Last 5 Advanced')

# Fetch and process the second set of data
data2 = fetch_nba_data(url, headers, params2)
df_filtered2 = process_data(data2, min_minutes=20, min_games=8)
df_ranked2 = rank_players(df_filtered2, rank_columns)
save_to_excel(df_ranked2, '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx', 'Last 10 Advanced')

data3 = fetch_nba_data(url, headers, params3)
df_filtered3 = process_data(data3, min_minutes=20, min_games=4)
df_ranked3 = rank_players(df_filtered3, rank_columns2)
save_to_excel(df_ranked3, '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx', 'Last 5 Basic')

# Fetch and process the second set of data
data4 = fetch_nba_data(url, headers, params4)
df_filtered4 = process_data(data4, min_minutes=20, min_games=8)
df_ranked4 = rank_players(df_filtered4, rank_columns2)
save_to_excel(df_ranked4, '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx', 'Last 10 Basic')


# Initialize an empty DataFrame
final_df = pd.DataFrame()

playtypes = ['OffScreen', 'Isolation', 'Transition', 'Postup', 'Spotup',
             'Handoff', 'Cut', 'OffRebound', 'Misc', 'PRRollman', 'PRBallHandler']

for x in playtypes:
    params5 = {
        "LeagueID": "00",
        "PerMode": "Totals",
        "PlayType": x,
        "PlayerOrTeam": "P",
        "SeasonType": "Regular Season",
        "SeasonYear": "2023-24",
        "TypeGrouping": "offensive"
    }
    data5 = fetch_nba_data(url2, headers, params5)
    headers_df = data5['resultSets'][0]['headers']
    rows = data5['resultSets'][0]['rowSet']
    df5 = pd.DataFrame(rows, columns=headers_df)
    df_filtered5 = df5[(df5['POSS'] >= 20)]

    # Calculate ranks for specified fields and append them to the dataframe
    fields_to_rank = [
        "POSS_PCT", "PPP", "FG_PCT", "FT_POSS_PCT", "TOV_POSS_PCT", "SF_POSS_PCT",
        "PLUSONE_POSS_PCT", "SCORE_POSS_PCT", "EFG_PCT", "POSS", "PTS",
        "FGM", "FGA", "FGMX"
    ]

    for field in fields_to_rank:
        rank_column_name = f"{field}_RANK"
        df_filtered5[rank_column_name] = df_filtered5[field].rank(method='min', ascending=False)

    # Append the data to the final DataFrame
    final_df = pd.concat([final_df, df_filtered5], ignore_index=True)

# Save the final DataFrame to an Excel file
excel_filename = '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx'
save_to_excel(final_df, '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx', 'Play Type')

generate_summary_tab(excel_filename)

print("Data processing and saving completed.")
