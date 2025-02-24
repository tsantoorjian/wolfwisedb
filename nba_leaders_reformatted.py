import requests
import pandas as pd
from openpyxl import load_workbook
import os

EXCEL_FILENAME = '/Users/tonysantoorjian/Documents/nba_player_stats.xlsx'
NBA_STATS_URL = "https://stats.nba.com/stats/leaguedashplayerstats"
NBA_SYNERGY_URL = 'https://stats.nba.com/stats/synergyplaytypes'
HEADERS = {
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9,pt-BR;q=0.8,pt;q=0.7",
    "Connection": "keep-alive",
    "Host": "stats.nba.com",
    "Origin": "https://www.nba.com",
    "Referer": "https://www.nba.com/",
    "Sec-Ch-Ua": "\"Not A Brand\";v=\"99\", \"Google Chrome\";v=\"121\", \"Chromium\";v=\"121\"",
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": "\"macOS\"",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
}
# Common rank columns
rank_columns = [
    "GP_RANK", "W_RANK", "L_RANK", "W_PCT_RANK", "MIN_RANK",
    "E_OFF_RATING_RANK", "OFF_RATING_RANK", "sp_work_OFF_RATING_RANK",
    "E_DEF_RATING_RANK", "DEF_RATING_RANK", "sp_work_DEF_RATING_RANK",
    "E_NET_RATING_RANK", "NET_RATING_RANK", "sp_work_NET_RATING_RANK",
    "AST_PCT_RANK", "AST_TO_RANK", "AST_RATIO_RANK", "OREB_PCT_RANK",
    "DREB_PCT_RANK", "REB_PCT_RANK", "TM_TOV_PCT_RANK", "E_TOV_PCT_RANK",
    "EFG_PCT_RANK", "TS_PCT_RANK", "USG_PCT_RANK", "E_USG_PCT_RANK",
    "E_PACE_RANK", "PACE_RANK", "sp_work_PACE_RANK", "PIE_RANK",
    "FGM_RANK", "FGA_RANK", "FGM_PG_RANK", "FGA_PG_RANK", "FG_PCT_RANK"
]

rank_columns2 = [
    "FGM_RANK", "FGA_RANK", "FG_PCT_RANK", "FG3M_RANK", "FG3A_RANK",
    "FG3_PCT_RANK", "FTM_RANK", "FTA_RANK", "FT_PCT_RANK", "OREB_RANK",
    "DREB_RANK", "REB_RANK", "AST_RANK", "TOV_RANK", "STL_RANK",
    "BLK_RANK", "BLKA_RANK", "PF_RANK", "PFD_RANK", "PTS_RANK",
    "PLUS_MINUS_RANK", "NBA_FANTASY_PTS_RANK", "DD2_RANK", "TD3_RANK",
    "WNBA_FANTASY_PTS_RANK"
]

rank_columns3 = [
    'SEASON_ID', 'PLAYER_ID', 'PLAYER_NAME', 'TEAM_ID',
    'TEAM_ABBREVIATION', 'TEAM_NAME', 'PLAY_TYPE', 'TYPE_GROUPING',
    'PERCENTILE', 'GP', 'POSS_PCT', 'PPP', 'FG_PCT', 'FT_POSS_PCT',
    'TOV_POSS_PCT', 'SF_POSS_PCT', 'PLUSONE_POSS_PCT', 'SCORE_POSS_PCT',
    'EFG_PCT', 'POSS', 'PTS', 'FGM', 'FGA', 'FGMX'
]

def get_params(stats_type, last_n_games=None, play_type=None):
    # Base parameters for regular stats
    params1 = {
        "College": "", "Conference": "", "Country": "", "DateFrom": "", "DateTo": "",
        "Division": "", "DraftPick": "", "DraftYear": "", "GameScope": "", "GameSegment": "",
        "Height": "", "ISTRound": "", "LastNGames": last_n_games, "LeagueID": "00",
        "Location": "", "MeasureType": "Advanced", "Month": "0", "OpponentTeamID": "0",
        "Outcome": "", "PORound": "0", "PaceAdjust": "N", "PerMode": "Totals",
        "Period": "0", "PlayerExperience": "", "PlayerPosition": "", "PlusMinus": "N",
        "Rank": "N", "Season": "2024-25", "SeasonSegment": "", "SeasonType": "Regular Season",
        "ShotClockRange": "", "StarterBench": "", "TeamID": "0", "VsConference": "",
        "VsDivision": "", "Weight": ""
    }

    # Base parameters for playtype stats
    params5 = {
        "LeagueID": "00", "PerMode": "Totals", "PlayType": play_type,
        "PlayerOrTeam": "P", "SeasonType": "Regular Season",
        "SeasonYear": "2024-25", "TypeGrouping": "offensive"
    }

    # Choose which parameter set to use
    if stats_type == 'regular':
        return params1
    elif stats_type == 'playtype':
        return params5
    else:
        raise ValueError("Invalid stats type specified.")

def process_data(data, filter_criteria):
    headers = data['resultSets'][0]['headers']
    rows = data['resultSets'][0]['rowSet']
    df = pd.DataFrame(rows, columns=headers)
    for column, value in filter_criteria.items():
        if column in df.columns:
            df = df[df[column] >= value]
    return df

def rank_players(df, rank_columns, fields_to_rank=None):
    if fields_to_rank:
        for field in fields_to_rank:
            rank_column_name = f"{field}_RANK"
            # Calculate rank and create/update the rank column
            df[rank_column_name] = df[field].rank(method='min', ascending=False)
    else:
        # Use existing rank columns for regular stats
        for col in rank_columns:
            # This assumes the "_RANK" columns already exist in the DataFrame
            # If additional logic is needed to handle cases where they don't, it can be added here
            df[col] = df[col].rank(method='average')
    return df


def fetch_nba_data(url, headers, params):
    print(f"Making API call to: {url}")
    print(f"With headers: {headers}")
    print(f"And parameters: {params}")

    response = requests.get(url, headers=headers, params=params)
    print(f"Response status code: {response.status_code}")

    if response.status_code == 200:
        try:
            return response.json()
        except ValueError as e:  # JSON decoding error
            print(f"Error decoding JSON: {e}")
            return None
    else:
        print(f"Failed to fetch data: {response.status_code}, Response: {response.text}")
        return None


def generate_summary_tab(excel_filename):
    if not os.path.exists(excel_filename):
        print(f"Error: The file {excel_filename} does not exist.")
        return

    # Initialize a DataFrame to collect the summary
    summary_df = pd.DataFrame()

    # Load the workbook
    wb = load_workbook(excel_filename)

    # Define a list to collect DataFrame pieces for the summary
    df_pieces = []

    # Prepare to filter sheets that contain Timberwolves data
    timberwolves_sheets = [sheet for sheet in wb.sheetnames if 'Rank' in sheet]

    # Loop through each Timberwolves sheet and collect the data
    for sheet_name in timberwolves_sheets:
        df = pd.read_excel(excel_filename, sheet_name=sheet_name)
        # Assuming 'Player' and the rank fields exist in the df
        # We filter the DataFrame for top 10 ranks
        for col in df.columns:
            if '_RANK' in col:
                # Extract the statistic's name by removing '_RANK' from the column name
                stat_name = col.replace('_RANK', '')
                top_ranks = df[df[col] <= 10][['PLAYER_NAME', col]]
                top_ranks['Stat'] = stat_name  # Add the statistic's name as 'Stat'
                top_ranks.rename(columns={col: 'Rank'}, inplace=True)  # Rename the rank column to 'Rank'
                top_ranks['Timeframe'] = sheet_name.replace(' Timberwolves',
                                                            '')  # Remove ' Timberwolves' to have a cleaner 'Timeframe'
                df_pieces.append(top_ranks)

    # Concatenate all the pieces to create the summary DataFrame
    if df_pieces:
        summary_df = pd.concat(df_pieces).sort_values(by=['PLAYER_NAME', 'Stat', 'Rank', 'Timeframe'])

        # Save the summary DataFrame to the 'Timberwolves Summary' sheet
        with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            summary_df.to_excel(writer, sheet_name='Timberwolves Summary', index=False)
        print("Timberwolves Summary generated successfully.")
    else:
        print("No top 10 ranks found.")

def save_to_excel(df, excel_filename, sheet_name):
    # Save the DataFrame to the Excel file
    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Load the workbook and move the 'Timberwolves Summary' sheet to the front
    workbook = load_workbook(excel_filename)
    sheet = workbook[sheet_name]  # Get the 'Timberwolves Summary' sheet
    sheet_idx = workbook.sheetnames.index(sheet_name)  # Get the index of the sheet
    workbook._sheets = [sheet] + workbook._sheets[:sheet_idx] + workbook._sheets[sheet_idx+1:]  # Move the sheet

    # Save the workbook
    workbook.save(excel_filename)
    print(f"'{sheet_name}' sheet moved to the front in '{excel_filename}'.")


def main():
    # Define URLs for fetching data
    regular_stats_url = "https://stats.nba.com/stats/leaguedashplayerstats"
    playtype_stats_url = "https://stats.nba.com/stats/synergyplaytypes"
    # List of Timberwolves players
    timberwolves_players = [
        "Julius Randle", "Rudy Gobert", "Mike Conley", "Jaden McDaniels",
        "Anthony Edwards", "Naz Reid", "Nickeil Alexander-Walker",
        "Donte DiVincenzo", "Joe Ingles", "Terrence Shannon Jr.", "Rob Dillingham", "Josh Minott"
    ]

    # DataFrame to collect all top 10 ranks
    top_10_ranks_df_pieces = []

    # Regular stats processing
    for measure_type in ["Advanced", "Base"]:
        for last_n_games in ['5', '10']:
            params = get_params('regular', last_n_games=last_n_games)
            params["MeasureType"] = measure_type

            data = fetch_nba_data(regular_stats_url, HEADERS, params)
            if data:
                gp_threshold = 4 if last_n_games == '5' else 8
                df_filtered = process_data(data, {'MIN': 20, 'GP': gp_threshold})

                # Select rank columns based on MeasureType
                rank_columns_to_use = rank_columns if measure_type == "Advanced" else rank_columns2
                df_ranked = rank_players(df_filtered, rank_columns_to_use)

                # Filter for Timberwolves players and extract top 10 ranks
                df_timberwolves_top_ranks = df_ranked[df_ranked['PLAYER_NAME'].isin(timberwolves_players)]

                # Iterate over the ranking columns and capture the rank and the corresponding value
                for col in rank_columns_to_use:
                    if '_RANK' in col:
                        stat_name = col.replace('_RANK', '')
                        top_ranks = df_timberwolves_top_ranks[
                            df_timberwolves_top_ranks[col] <= 10
                            ][['PLAYER_NAME', col, stat_name]].copy()  # Capture both rank and stat value
                        top_ranks.rename(columns={col: 'Rank', stat_name: 'Stat_Value'}, inplace=True)
                        top_ranks['Stat'] = stat_name
                        top_ranks['Timeframe'] = f'Last {last_n_games} {measure_type}'
                        top_10_ranks_df_pieces.append(top_ranks)

    # Play type stats processing
    play_types = ['OffScreen', 'Isolation', 'Transition', 'Postup', 'Spotup',
                  'Handoff', 'Cut', 'OffRebound', 'Misc', 'PRRollman', 'PRBallHandler']
    fields_to_rank_playtypes = [
        "POSS_PCT", "PPP", "FG_PCT", "FT_POSS_PCT", "TOV_POSS_PCT", "SF_POSS_PCT",
        "PLUSONE_POSS_PCT", "SCORE_POSS_PCT", "EFG_PCT", "POSS", "PTS", "FGM", "FGA", "FGMX"
    ]

    for play_type in play_types:
        params = get_params('playtype', play_type=play_type)
        data = fetch_nba_data(playtype_stats_url, HEADERS, params)
        if data:
            df_filtered = process_data(data, {'POSS': 20})
            df_ranked = rank_players(df_filtered, rank_columns=[], fields_to_rank=fields_to_rank_playtypes)

            # Filter for Timberwolves players and extract top 10 ranks
            df_timberwolves_top_ranks = df_ranked[df_ranked['PLAYER_NAME'].isin(timberwolves_players)]

            # Iterate over the fields to rank and capture the rank and the corresponding value
            for field in fields_to_rank_playtypes:
                rank_column_name = f"{field}_RANK"
                top_ranks = df_timberwolves_top_ranks[
                    df_timberwolves_top_ranks[rank_column_name] <= 10
                ][['PLAYER_NAME', rank_column_name, field]].copy()  # Capture both rank and stat value
                top_ranks.rename(columns={rank_column_name: 'Rank', field: 'Stat_Value'}, inplace=True)
                top_ranks['Stat'] = field
                top_ranks['Timeframe'] = f'Playtype - {play_type}'
                top_10_ranks_df_pieces.append(top_ranks)

    # Concatenate all top 10 ranks DataFrame pieces
    if top_10_ranks_df_pieces:
        summary_df = pd.concat(top_10_ranks_df_pieces).sort_values(
            by=['PLAYER_NAME', 'Stat', 'Rank', 'Timeframe']
        )
        summary_df = summary_df[['PLAYER_NAME', 'Stat', 'Stat_Value', 'Rank', 'Timeframe']]  # Reorder columns
        # Save the summary DataFrame to the 'Timberwolves Summary' sheet
        save_to_excel(summary_df, EXCEL_FILENAME, 'Timberwolves Summary')
    else:
        print("No top 10 ranks found.")

    print("Data processing, saving, and summary generation completed.")


if __name__ == "__main__":
    main()
