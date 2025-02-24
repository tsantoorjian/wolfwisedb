import requests
import pandas as pd
from openpyxl import load_workbook




def fetch_nba_data(url, headers, params):
    response = requests.get(url, headers=headers, params=params)
    data = response.json()
    return data

def process_data(data, min_attempts):
    headers = data['resultSets'][0]['headers']
    rows = data['resultSets'][0]['rowSet']
    df = pd.DataFrame(rows, columns=headers)
    df_filtered = df[(df['FG3A'] >= min_attempts)]
    return df_filtered

def save_to_excel(df, excel_filename, sheet_name):
    # Load the existing workbook
    workbook = load_workbook(excel_filename)

    # Check if the sheet exists and delete it
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    # Now, use pandas ExcelWriter with the updated workbook
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save the workbook
    workbook.save(excel_filename)


# Assuming other parts of your script have already been defined, such as fetching and processing data

def generate_summary_tab(excel_filename):
    # Load the workbook and check if 'Timberwolves Summary' sheet exists
    wb = load_workbook(excel_filename)
    if 'Timberwolves Summary' in wb.sheetnames:
        del wb['Timberwolves Summary']  # Delete the sheet if it exists
        wb.save(excel_filename)  # Save the workbook after deletion

    sheet_names = wb.sheetnames
    rank_sheets = [name for name in sheet_names if 'Rank' in name and name != 'Timberwolves Summary']

    top_10_ranks_with_timeframe = []

    for sheet in rank_sheets:
        # Specifying header=1 to indicate that the second row contains the column names
        df = pd.read_excel(excel_filename, sheet_name=sheet, header=1)
        # Adjusting based on the actual structure of your sheets
        if 'Player' in df.columns:
            for col in df.columns[2:]:  # Assuming the first two columns are not needed for rank calculation
                df[col] = pd.to_numeric(df[col], errors='coerce')
                top_ranks = df.loc[df[col] <= 10, ['Player', col]]
                for _, row in top_ranks.iterrows():
                    top_10_ranks_with_timeframe.append({
                        "Player": row['Player'],
                        "Stat": col,
                        "Rank": row[col],
                        "Timeframe": sheet
                    })

    # Convert the list to DataFrame, sort, and save
    if top_10_ranks_with_timeframe:
        summary_df = pd.DataFrame(top_10_ranks_with_timeframe).sort_values(by="Player")
        save_to_excel(summary_df, excel_filename, 'Timberwolves Summary')
    else:
        print("No top 10 ranks found.")

# Common headers and URL
url = "https://stats.nba.com/stats/boxscoretraditionalv3"
headers = {
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Accept-Language": "en-US,en;q=0.9,pt-BR;q=0.8,pt;q=0.7",
    "Connection": "keep-alive",
    "Host": "stats.nba.com",
    "Origin": "https://www.nba.com",
    "Referer": "https://www.nba.com/",
    "Sec-Ch-Ua": "\"Chromium\";v=\"122\", \"Not A Brand\";v=\"24\", \"Google Chrome\";v=\"122\"",
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": "\"macOS\"",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
}

params = {
    "GameID": "0022300873",
    "LeagueID": "00",
    "endPeriod": "4",
    "endRange": "28800",
    "rangeType": "1",
    "startPeriod": "1",
    "startRange": "0",
}



# Fetch and process the first set of data
data1 = fetch_nba_data(url, headers, params)
save_to_excel(data1, '/Users/tonysantoorjian/Documents/game_log_stats.xlsx', 'Sheet1')
